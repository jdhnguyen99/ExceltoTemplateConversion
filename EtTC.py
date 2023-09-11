import pandas as pd
import openpyxl
import os
import shutil
import tkinter as tk
import re
from tkinter import filedialog, StringVar
from docxtpl import DocxTemplate
from docxcompose.composer import Composer
from docx import Document as Document_compose

# UI Design
window = tk.Tk()
window.geometry('750x250')
window.title("Excel To Template Converter")

window.rowconfigure(0, minsize=5)
window.columnconfigure(0, weight=1)
window.columnconfigure(1, weight=1)
window.columnconfigure(2, weight=1)

# All Labels
label1 = tk.Label(text = "Excel to Template Converter", bg = "white", fg = "black")
label1.grid(row = 0, column = 1)
label2 = tk.Label(text = "Input Excel File:", bg = "white", fg = "black")
label2.grid(row = 1, column = 0)
label3 = tk.Label(text = "Output Folder Location of Template Doc:", bg = "white", fg = "black")
label3.grid(row = 2, column = 0)
labelFinish = tk.Label(text = "Ready to Generate.",  fg = "black", font = "Calibri 12 bold")
labelFinish.grid(row = 4, column = 1)

# All entries for file path
excel_path = StringVar()
word_path = StringVar()
entry1 = tk.Entry(bg = "light grey",textvariable = excel_path, width = 50, fg = "black", highlightthickness = 2)
entry1.grid(row = 1, column = 1)
entry2 = tk.Entry(bg = "light grey",textvariable = word_path, width = 50, fg = "black", highlightthickness = 2)
entry2.grid(row = 2, column = 1)

def handle_mouseclickExcel():
    excel_path.set(filedialog.askopenfilename())

def handle_mouseclickWord():
    word_path.set(filedialog.askdirectory())

def unmerge_cells(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    merged_cells = ws.merged_cells
    merged_ranges = []

    for merged_cell_range in merged_cells:
        merged_ranges.append(merged_cell_range.coord)
    
    for merged_range in merged_ranges:
        ws.unmerge_cells(merged_range)

    wb.save(excel_path)
    wb.close

def combine_all_docx(filename_master, files_list):
    number_of_selections = len(files_list)
    master = Document_compose(filename_master)
    composer = Composer(master)
    for i in range(0, number_of_selections):
        doc_temp = Document_compose(files_list[i])
        composer.append(doc_temp)
    composer.save(f"{word_path.get()}\Combined Pt Info Document.docx")

def close_window():
    window.destroy()

def handle_mouseclickGenerate():
    try:
        unmerge_cells(excel_path.get())
        df = pd.read_excel(excel_path.get(), usecols = ['Patient First Name', 'Patient Last Name', 'Gender', 'MRN', 'DOB', 'Address 1', 'Address 2', 'City', 'State', 'ZipCode', 'Phone', 'Appt Date', 'Appt Time', 'Location', 'Provider', 'Notes']) 
    except ValueError as e:
        labelFinish.config(text = f"""These columns are either missing or misspelled: {', '.join(re.findall(r"'(.*?)'", e.args[0]))}""", fg = "dark red")
        return
    except PermissionError as e:
        labelFinish.config(text = "Scheduling excel is currently opened.\nPlease close the file and click generate again", fg = "dark red")
        return
    
    df = df.fillna('')

    
    rows_to_remove = []
    consolidated_note_index = 0
    for i, row in df.iterrows():
        if i > 0 and row['Patient First Name'] != '':
            consolidated_note_index = i
        if i > 0 and row['Patient First Name'] == '' and row['Notes'] != '':
            df.at[consolidated_note_index, 'Notes'] += '\n' + row['Notes']
            rows_to_remove.append(i)
    df = df[df['Notes'] != '']
    df = df.drop(rows_to_remove)
    df.reset_index(drop=True, inplace=True)
    
    
    doc = DocxTemplate(f'{word_path.get()}\PtInfoTemplate.docx')               # open doc for write
    filepath = "tempJDHN"

    try: 
        open(f'{word_path.get()}\Combined Pt Info Document.docx', 'w')
    except PermissionError as e:
        labelFinish.config(text = "Combined Pt Info Document.docx is currently opened.\nPlease close the file and click generate again", fg = "dark red")
        return
    
    if os.path.exists(f"{word_path.get()}\{filepath}"):                        # delete temp folder if it exists
        shutil.rmtree(f"{word_path.get()}\{filepath}")                             
    
    os.mkdir(f"{word_path.get()}\{filepath}")
    for index, row in df.iterrows():                                           # Iterate through the rows and create contexts for each entry
        context = {
            "Patient_First_Name": row["Patient First Name"],
            "Patient_Last_Name": row["Patient Last Name"],
            "Gender": row["Gender"],
            "MRN": row["MRN"],
            "DOB": row["DOB"].strftime("%m-%d-%Y"),
            "Address_1": row["Address 1"],
            "Address_2": row["Address 2"],
            "City": row["City"],
            "State": row["State"],
            "ZipCode": row["ZipCode"],
            "Phone": row["Phone"],
            "Appt_Date": row["Appt Date"].strftime("%m-%d-%Y"),
            "Appt_Time": row["Appt Time"].strftime("%I:%M %p"),
            "Location": row["Location"],
            "Provider": row["Provider"],
            "Notes": row["Notes"]
        }
        doc.render(context)
        doc.save(f'{word_path.get()}\{filepath}\Pt Info Document{index}.docx')

    filename_master = f"{word_path.get()}\{filepath}\Pt Info Document0.docx"
    files_list = [f'{word_path.get()}\{filepath}\Pt Info Document{index}.docx' for index in range(len(df))]
    combine_all_docx(filename_master, files_list)
    shutil.rmtree(f"{word_path.get()}\{filepath}")
    labelFinish.config(text = "Combined Pt Info Document.docx has been created!", fg = "dark green")
    button4 = tk.Button(text = "OK", command = close_window, width = 5, height = 3, bg = "grey", fg = "black")
    button4.grid(row = 5, column = 1)



# All Buttons
button1 = tk.Button(text = "Browse",command = handle_mouseclickExcel, bg = "light blue", fg = "black")
button1.grid(row = 1, column = 2)
button2 = tk.Button(text = "Browse",command = handle_mouseclickWord, bg = "light blue", fg = "black")
button2.grid(row = 2, column = 2)
button3 = tk.Button(text = "GENERATE",command = handle_mouseclickGenerate, bg = "light grey", fg = "black")
button3.grid(row = 3, column = 1)

window.mainloop()

################# NOTES ###################
# 1. pyinstaller --onefile -w --hidden-import openpyxl.cell._writer 
# --collect-data "docxcompose"  EtTC.py
# 2. Remember to create a new environment to only include necessary
#    packages
# 3. Remember to check for correct pyinstaller versions
###########################################